from openpyxl import load_workbook

file_path = 'sequences.xlsx'
marker_sheet = '6_8'

wb = load_workbook(file_path)
sheet_names = wb.sheetnames

if marker_sheet not in sheet_names:
    raise ValueError(f"Sheet '{marker_sheet}' not found in workbook.")

marker_index = sheet_names.index(marker_sheet)
sheets_to_delete = sheet_names[marker_index + 1:]

for sheet in sheets_to_delete:
    print(f"Deleting sheet: {sheet}")
    del wb[sheet]

wb.save(file_path)
print("All sheets after '6_8' have been deleted.")
