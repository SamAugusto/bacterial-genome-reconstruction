import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = r'C:\Users\Samuel\OneDrive - Drexel University\Dr. Xiao Scripts Coop 2025\kmer\pacbio_sequences.xlsx'
prefixes = [
    '1_2', '2_3', '2_4',
    '3_3', '3_4', '3_5',
    '4_4', '4_5', '4_6',
    '5_4', '5_5', '5_6', '5_7',
    '6_6', '6_7', '6_8',
    '7_6', '7_7', '7_8', '7_9',
    '8_8', '8_9', '8_10',
    '9_9', '9_10', '9_11',
    '10_10', '10_11', '10_12',
    '11_10', '11_11',
    '12_11', '13_12'
]

wb = load_workbook(file_path)
sheet_names = wb.sheetnames

results = []

for prefix in prefixes:
    matching_sheets = [name for name in sheet_names if name.startswith(prefix + '__') and name != prefix]
    
    for sheet_name in matching_sheets:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        min_hits_sum = df.groupby('Kmer_seq')['hits'].min().sum()
        results.append({'Sheet Name': sheet_name, 'Min Hits Sum': min_hits_sum})
        ws = wb[sheet_name]
        last_col = get_column_letter(ws.max_column + 1)
        ws[f'{last_col}1'] = min_hits_sum

wb.save(file_path)

summary_df = pd.DataFrame(results)
summary_df.to_excel('min_hits_summary.xlsx', index=False)
print("✅ All sheets processed. Summary saved to 'test_min_hits_summary.xlsx'.")
