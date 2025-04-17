import pandas as pd
from openpyxl import load_workbook

file_path = 'sequences.xlsx'

# Base sheet prefixes you want to include (like '2_4', '3_4', etc.)
base_prefixes = ['1_2','2_3','2_4','3_4','3_5','4_4','4_5','4_6','5_5','5_6','5_7','6_6','6_8']

wb = load_workbook(file_path)
sheet_names = wb.sheetnames

results = []

# Loop through all sheets
for sheet_name in sheet_names:
    # Only process sheets that start with one of the base prefixes followed by "__"
    if not any(sheet_name.startswith(prefix + '__') for prefix in base_prefixes):
        continue

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if 'Kmer_seq' in df.columns and 'hits' in df.columns:
            grouped = df.groupby('Kmer_seq')['hits']
            min_hits_sum = grouped.min().sum()
            unique_kmers = grouped.ngroups
            results.append({
                'Pattern': sheet_name,
                'sum_minim_hit': min_hits_sum,
                'unique_kmers': unique_kmers
            })
    except Exception as e:
        print(f"⚠️ Skipping sheet '{sheet_name}' due to error: {e}")

# Convert to DataFrame and save to Excel
summary_df = pd.DataFrame(results)
summary_df.to_excel('final_min_hit_summary_all.xlsx', index=False)
print("✅ Done! Summary with all prefix combo sheets written to 'final_min_hit_summary_all.xlsx'")
