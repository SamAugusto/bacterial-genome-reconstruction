from openpyxl import load_workbook

# === Setup ===
file_path = 'sequences.xlsx'
marker_sheet = '6_8'

# Load the workbook
wb = load_workbook(file_path)

# Get list of all sheet names
sheet_names = wb.sheetnames

# Find the position of the marker sheet
if marker_sheet not in sheet_names:
    raise ValueError(f"Sheet '{marker_sheet}' not found in workbook.")

marker_index = sheet_names.index(marker_sheet)

# Sheets after the marker
sheets_to_delete = sheet_names[marker_index + 1:]

# Delete them
for sheet in sheets_to_delete:
    print(f"🗑️ Deleting sheet: {sheet}")
    del wb[sheet]

# Save the workbook
wb.save(file_path)

print("✅ All sheets after '6_8' have been deleted.")
