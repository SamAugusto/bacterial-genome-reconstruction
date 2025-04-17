import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_path = 'sequences.xlsx'
prefixes = ['1_2', '2_3', '2_4' '3_4', '3_5', '4_4', '4_5', '4_6', '5_5', '5_6', '5_7', '6_6', '6_8']

wb = load_workbook(file_path)
sheet_names = wb.sheetnames

results = []

for prefix in prefixes:
    matching_sheets = [name for name in sheet_names if name.startswith(prefix + '__') and name != prefix]
    
    for sheet_name in matching_sheets:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        min_hits_sum = df.groupby('Kmer_seq')['hits'].min().sum()
        results.append({'Sheet Name': sheet_name, 'Min Hits Sum': min_hits_sum})
        ws = wb[sheet_name]
        last_col = get_column_letter(ws.max_column + 1)
        ws[f'{last_col}1'] = min_hits_sum

wb.save(file_path)

summary_df = pd.DataFrame(results)
summary_df.to_excel('min_hits_summary.xlsx', index=False)
print("✅ All sheets processed. Summary saved to 'min_hits_summary.xlsx'.")
